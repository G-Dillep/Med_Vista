from flask import Flask, request, jsonify, render_template,flash,redirect,url_for
import pandas as pd 
from datetime import datetime
from openpyxl import load_workbook
import os
from apscheduler.schedulers.background import BackgroundScheduler
from twilio.rest import Client
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS
import google.generativeai as genai
import uuid
import csv,enum
from datetime import timedelta,date,datetime

app = Flask(__name__)
# Load the CSV data for medicines
def load_csv_data():
    try:
        df = pd.read_csv("medicine_dataset_main.csv", low_memory=False)
        return df
    except FileNotFoundError:
        print("Error: CSV file not found")
        return None
    except pd.errors.EmptyDataError:
        print("Error: CSV file is empty")
        return None

df = load_csv_data()
# ==================================================================================
@app.route("/")
def index():
    if df is None:
        return "Error: Unable to load CSV data"
    return render_template("start.html")
# ===================================================================================
@app.route("/homepage")
def homepage():
    return render_template("homepage.html")
@app.route("/guesthomepage")
def guesthomepage():
    return render_template("guesthomepage.html")
# ====================================================================================
def get_first_non_null(row, fields):
    for field in fields:
        value = row.get(field)  # Safely access the field
        if pd.notnull(value) and value.strip() != "":  # Check for non-null, non-empty
            return value
    return "Not available"

@app.route("/search_for_medicine", methods=['GET', 'POST'])
def search_for_medicine():
    if request.method == "POST":
        search_type = request.form.get("search-type")
        search_term = request.form.get("search-input").lower()
        results = []
        
        if search_type == "name":
            filtered_df = df[df["name"].str.lower().str.contains(search_term, na=False)]
            
            for _, row in filtered_df.iterrows():
                result = {
                    "name": row.get("name", "Unknown"),
                    "use1": row.get("use0", "Not available"),  # Directly take from use0
                    "use2": get_first_non_null(row, ["use1", "use2", "use3"]),  # First non-null among use1, use2, use3
                    "substitute1": get_first_non_null(row, ["substitute0"]),
                    "substitute2": get_first_non_null(row, ["substitute1"]),
                }
                results.append(result)

        
        return jsonify(results)
    return render_template("search_for_medicine.html")

# ====================================================================================
# Path for saving contact forms
excel_path = 'contact_data.xlsx'

@app.route("/contactpage")
def contact():
    return render_template("contactpage.html")

@app.route('/submit_contact_form', methods=['POST'])
def submit_contact_form():
    try:
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        query = request.form.get('query')
        suggestions = request.form.get('suggestions')
        if not name or not email or not query:
            return jsonify({'message': 'Name, email, and query are required fields.'}), 400
        data = {
            'Name': [name],
            'Phone': [phone],
            'Email': [email],
            'Query': [query],
            'Suggestions': [suggestions],
            'Date Submitted': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        }
        df = pd.DataFrame(data)
        if not os.path.exists(excel_path):
            df.to_excel(excel_path, index=False)
        else:
            with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                workbook = load_workbook(excel_path)
                sheet = workbook.active
                startrow = sheet.max_row
                df.to_excel(writer, index=False, header=False, startrow=startrow)
        return jsonify({'message': 'Your contact information has been successfully saved. Thank you!'}), 200
    except Exception as e:
        print(f"Error while saving data: {e}")
        return jsonify({'message': 'Failed to save data. Please try again later.'}), 500

# ====================================================================================
def get_dates_in_range(start_date, end_date, days_of_week):
    """Generates all dates within a range that match selected days of the week."""
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    delta = timedelta(days=1)
    selected_dates = []

    while start <= end:
        if start.strftime("%A") in days_of_week:
            selected_dates.append(start.strftime("%Y-%m-%d"))
        start += delta
    return selected_dates

# Routes
def append_to_csv(file_path, data, fieldnames):
    """Appends a single row of data to a CSV file."""
    with open(file_path, 'a', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writerow(data)
@app.route('/signup_doctor', methods=['GET', 'POST'])
def signup_doctor():
    if request.method == 'POST':
        doctor_id =request.form['dr_id']   # Generate a unique doctor ID
        name = request.form['username']
        email = request.form['email']
        password = request.form['password']
        mobile_number = request.form['mobile_number']
        specialization = request.form['specialization']
        location = request.form['location']
        geolocation = request.form['geolocation']

        # Save doctor information to doctors.csv
        doctor_data = {
            'doctor_id': doctor_id,
            'name': name,
            'email': email,
            'password': generate_password_hash(request.form['password']),  # Ideally, hash this password
            'mobile_number': mobile_number,
            'specialization': specialization,
            'location': location,
            'geolocation': geolocation
        }
        doctor_fieldnames = ['doctor_id', 'name', 'specialization', 'location', 'email', 'password', 'mobile_number', 'geolocation']
        append_to_csv('doctors.csv', doctor_data, doctor_fieldnames)

        # Collect availability information
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        days_of_week = request.form.getlist('days')
        start_times = request.form.getlist('start_time[]')
        end_times = request.form.getlist('end_time[]')

        # Generate and save availability slots
        availability_fieldnames = ['doctor_id', 'date', 'time_slot', 'is_available']
        dates = get_dates_in_range(start_date, end_date, days_of_week)

        # Append time slots for each day within the range
        for date in dates:
            for start_time, end_time in zip(start_times, end_times):
                slot_data = {
                    'doctor_id': doctor_id,
                    'date': date,
                    'time_slot': f"{start_time}-{end_time}",
                    'is_available': 'TRUE'
                }
                append_to_csv('availability.csv', slot_data, availability_fieldnames)

        flash("Doctor signed up and availability slots added successfully!")
        return redirect(url_for('homepage'))

    return render_template('signup_doctor.html')
#==================================================================================================================================================================================================================================================================================================

@app.route('/search_for_doctor', methods=['GET', 'POST'])
def search_for_doctor():
    file_path = 'doctors.csv'
    doctors_df = pd.read_csv(file_path)
    specializations = doctors_df['specialization'].dropna().unique()
    locations = doctors_df['location'].dropna().unique()
    if request.method == 'POST':
        specialization = request.form['specialization']
        location = request.form['location']
        return redirect(url_for('search_results', specialization=specialization, location=location))

    return render_template('search_for_doctor.html',specializations=specializations,locations=locations)

@app.route('/search-results')
def search_results():
    specialization = request.args.get('specialization', '').strip()
    location = request.args.get('location', '').strip()

    # Load the CSV file
    file_path = 'doctors.csv'
    doctors_df = pd.read_csv(file_path)

    # Filter doctors based on specialization and location
    filtered_doctors = doctors_df[
        (doctors_df['specialization'].str.contains(specialization, case=False, na=False)) &
        (doctors_df['location'].str.contains(location, case=False, na=False))
    ]
    # Convert the filtered data to a list of dictionaries
    doctors = filtered_doctors.to_dict(orient='records')

    return render_template('search_results.html', doctors=doctors)

@app.route('/doctor-profile/<doctor_id>')
def doctor_profile(doctor_id):
    # Load the CSV file
    file_path = 'doctors.csv'
    doctors_df = pd.read_csv(file_path)
    # Filter the doctor by ID
    doctor_data = doctors_df[doctors_df['doctor_id'] == int(doctor_id)]
    # If doctor not found, return 404
    if doctor_data.empty:
        app.logger.error(f"Doctor with ID {doctor_id} not found")
        return "Doctor not found", 404
    # Convert the doctor's data to a dictionary
    doctor = doctor_data.iloc[0].to_dict()

    return render_template('doctor_profile.html', doctor=doctor)


# Helper functions to read/write CSV files
def read_csv(file_name):
    """Reads CSV data into a list of dictionaries."""
    data = []
    if os.path.exists(file_name):
        with open(file_name, mode='r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                data.append(row)
    return data

def write_csv(file_name, data, fieldnames):
    """Writes a list of dictionaries to a CSV file."""
    with open(file_name, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

# availability', methods=['POST'])
@app.route('/check_availability', methods=['POST'])
def check_availability():
    # Parse the request data
    data = request.json
    doctor_id = int(data['doctor_id'])  # Convert doctor_id to integer
    date = data['date'].strip()  # Ensure no extra spaces in date

    # Load and normalize availability data from CSV
    availability_data = []
    with open('availability.csv', 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            availability_data.append({
                "doctor_id": int(row["doctor_id"]),  # Ensure doctor_id is integer
                "date": row["date"].strip(),  # Normalize date by stripping spaces
                "time_slot": row["time_slot"].strip(),  # Strip spaces in time_slot
                "is_available": row["is_available"].strip().upper() == "TRUE"  # Normalize is_available
            })

    # # Debug: Print the loaded availability data
    # print("Loaded Availability Data:", availability_data)

    # Filter available slots
    available_slots = [
        slot['time_slot'] for slot in availability_data
        if slot['doctor_id'] == doctor_id and slot['date'] == date and slot['is_available']
    ]
    return jsonify({"available_slots": available_slots})
@app.route('/book_appointment', methods=['POST'])
def book_appointment():
    data = request.json  # Expecting JSON data
    doctor_id = int(data.get('doctor_id'))  # Ensure doctor_id is an integer
    user_email = data.get('user_email')
    date = data.get('date')
    time_slot = data.get('time_slot')

    # Save appointment data to CSV
    with open('appointments.csv', 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([user_email, doctor_id, date, time_slot])

    # Read availability data and update the relevant slot
    updated_availability_data = []
    with open('availability.csv', 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            # Check if this is the slot to be updated
            if (
                int(row['doctor_id']) == doctor_id and
                row['date'] == date and
                row['time_slot'] == time_slot
            ):
                row['is_available'] = "FALSE"  # Mark as unavailable
            updated_availability_data.append(row)

    # Write updated data back to the CSV
    with open('availability.csv', 'w', newline='') as file:
        fieldnames = ['doctor_id', 'date', 'time_slot', 'is_available']
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_availability_data)

    return jsonify({"message": "Appointment booked successfully!"})


# ====================================================================================

TWILIO_SID = 'AC2dbfe4867c6bbc48c8d38c09480adecc'
TWILIO_AUTH_TOKEN = 'c205ba353a899631e3a7b3bced918d1a'
TWILIO_WHATSAPP_NUMBER = 'whatsapp:+14155238886'
client = Client(TWILIO_SID, TWILIO_AUTH_TOKEN)

scheduler = BackgroundScheduler()
scheduler.start()

csv_file_path = 'reminders.csv'

def ensure_csv_exists():
    if not os.path.exists(csv_file_path):
        pd.DataFrame(columns=['Phone', 'Message', 'Time']).to_csv(csv_file_path, index=False)
        os.chmod(csv_file_path, 0o666)

ensure_csv_exists()

@app.route("/remainderpage")
def remainderpage():
    return render_template("remainderpage.html")

@app.route('/schedule_reminder', methods=['POST'])
def schedule_reminder():
    phone_number = request.form['phone']
    reminder_message = request.form['message']
    reminder_time_str = request.form['time']

    try:
        reminder_time = datetime.strptime(reminder_time_str, '%Y-%m-%dT%H:%M')

        job_id = f'reminder_{phone_number}_{reminder_time}'
        scheduler.add_job(send_whatsapp_message, 'date', run_date=reminder_time,
                          args=[phone_number, reminder_message], id=job_id)

        reminder_data = {'Phone': phone_number, 'Message': reminder_message, 'Time': reminder_time_str}
        pd.DataFrame([reminder_data]).to_csv(csv_file_path, mode='a', index=False, header=not os.path.exists(csv_file_path))

        return jsonify({'message': 'Reminder scheduled successfully!'})
    except ValueError:
        return jsonify({'message': 'Incorrect date format. Please use YYYY-MM-DDTHH:MM.'}), 400
    except Exception as e:
        return jsonify({'message': f'Failed to schedule reminder: {e}'}), 500

@app.route('/view_reminders', methods=['GET'])
def view_reminders():
    try:
        if os.path.exists(csv_file_path):
            reminders = pd.read_csv(csv_file_path).to_dict(orient='records')
            return jsonify(reminders), 200
        return jsonify({'message': 'No reminders found.'}), 200
    except Exception as e:
        return jsonify({'message': f'Failed to load reminders: {e}'}), 500

def send_whatsapp_message(phone_number, reminder_message):
    try:
        formatted_number = f'whatsapp:+{phone_number.lstrip("+")}'
        message = client.messages.create(
            body=reminder_message,
            from_=TWILIO_WHATSAPP_NUMBER,
            to=formatted_number
        )
        print(f"WhatsApp message sent: {message.sid}")
    except Exception as e:
        print(f"Failed to send WhatsApp message: {e}")

# ====================================================================================

app.secret_key = 'reddy'
user_data_path = "user_data.csv"
doctor_data_path = "doctor_data.csv"

def ensure_csv_exists(file_path, columns):
    if not os.path.exists(file_path):
        pd.DataFrame(columns=columns).to_csv(file_path, index=False)

ensure_csv_exists(user_data_path, ['fullname','username','dob','mobile_number', 'email','address', 'password'])
@app.route("/signup_user", methods=["GET", "POST"])
def signup_user():
    if request.method == "POST":
        fullname=request.form['fullname']
        username = request.form['username']
        dob = request.form['dob']
        mobile_number = request.form['mobile_number']
        email = request.form['email']
        address = request.form['address']
        password = generate_password_hash(request.form['password'])

        user_df = pd.DataFrame([[fullname,username,dob,mobile_number, email,address, password]], columns=['fullname','username','dob', 'mobile_number','email','address', 'password'])
        user_df.to_csv(user_data_path, mode='a', header=False, index=False)

        flash("User signup successful!")
        return redirect(url_for('index'))
    
    return render_template("signup_user.html")

@app.route("/login_user", methods=["GET", "POST"])
def login_user():
    if request.method == "POST":
        email = request.form['email']
        password = request.form['password']

        # Verify user login
        user_df = pd.read_csv(user_data_path)
        user = user_df[user_df['email'] == email]
        if not user.empty and check_password_hash(user.iloc[0]['password'], password):
            flash("User login successful!")
            return redirect(url_for('homepage'))
        flash("Invalid credentials, please try again.")
    
    return render_template("login_user.html")


#=========================================================================================
def ensure_csv_exists(file_path, columns):
    # Create the CSV file if it doesn't exist
    if not os.path.exists(file_path):
        pd.DataFrame(columns=columns).to_csv(file_path, index=False)

# Ensure the Doctor CSV file exists

# @app.route("/onlineappointments")
# def onlineappointments():
#     return render_template("onlineappointments.html")


#====================================================================================


CORS(app)  # Allow cross-origin requests for frontend-backend communication

# Configure Generative AI
genai.configure(api_key="AIzaSyBQabu2ruJ3nTrP4IFSqD9TrGx10gMPNOM")

# Configure the model
generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    "response_mime_type": "text/plain",
}

# Initialize chat session
model = genai.GenerativeModel(
    model_name="gemini-1.5-pro",
    generation_config=generation_config,
    system_instruction="""You are an intelligent assistant designed exclusively for the MedExpert application.system_instruction You are an intelligent assistant designed exclusively for the MedExpert application. 
MedExpert provides the following services:
1. Detailed information on medicines, including their uses, side effects, and alternatives.
2. Suggestions based on a medicine database, such as Coolac and others.
3. Guidance on medicine timing and dosage (but no personalized prescriptions).
4. General health advice related to medicine use.

Rules:
- Provide concise responses (3–5 sentences maximum).
- Avoid unnecessary elaboration or unrelated details.
- Focus only on the user's question, directly answering it.
- Do not include unrelated advice or information.
"""
)
chat_session = model.start_chat()
@app.route('/chatbot')
def chatbot():
    """Serve the index HTML file."""
    return render_template('chatbot.html')

@app.route('/ask', methods=['POST'])
def ask_question():
    """Endpoint to process user questions."""
    data = request.json
    question = data.get('question', '')

    if not question:
        return jsonify({"answer": "Please provide a valid question."}), 400

    try:
        # Send question to the Generative AI model
        response = chat_session.send_message(question)
        return jsonify({"answer": response.text})
    except Exception as e:
        print(f"Error processing question: {e}")
        return jsonify({"answer": "There was an error processing your question."}), 500
    


#=============================================================================================================================================================================================================
doctor_data_path = "doctors.csv"

def ensure_csv_exists(file_path, columns):
    if not os.path.exists(file_path):
        pd.DataFrame(columns=columns).to_csv(file_path, index=False)




@app.route('/doctor-login', methods=['GET', 'POST'])
def doctor_login():
    if request.method == 'POST':
        email = request.form['username']
        password = request.form['password']
        user_df = pd.read_csv(doctor_data_path)
        user = user_df[user_df['email'] == email]
        if not user.empty and check_password_hash(user.iloc[0]['password'], password):
            flash("User login successful!")
            print("login ayyindi")
            return redirect(url_for('doctor_dashboard'))
        else:
            print("login avvale")
        flash("Invalid credentials, please try again.")

    return render_template('doctorlogin.html')

# Route: Doctor Dashboard
@app.route('/doctor-dashboard')
def doctor_dashboard():
    # Load availability data
    availability = []
    with open('availability.csv', 'r') as file:
        reader = csv.reader(file)
        availability = list(reader)

    # Load patient data
    patients = []
    with open('user_data.csv', 'r') as file:
        reader = csv.reader(file)
        patients = list(reader)

    # Enumerate availability data in Python
    enumerated_availability = list(enumerate(availability))

    return render_template('doctordashboard.html', availability=enumerated_availability, patients=patients)



# Route: Add Appointment Slot
@app.route('/add-slot', methods=['POST'])
def add_slot():
    date = request.form['date']
    time = request.form['time']
    with open('availability.csv', 'a') as file:
        writer = csv.writer(file)
        writer.writerow([date, time, 'Available'])
    flash("Appointment slot added successfully!")
    return redirect(url_for('doctor_dashboard'))

# Route: Delete Appointment Slot
@app.route('/delete-slot', methods=['POST'])
def delete_slot():
    slot_id = int(request.form['slot_id'])
    availability = []
    with open('availability.csv', 'r') as file:
        reader = csv.reader(file)
        availability = list(reader)
    if 0 <= slot_id < len(availability):
        availability.pop(slot_id)
        with open('availability.csv', 'w') as file:
            writer = csv.writer(file)
            writer.writerows(availability)
        flash("Appointment slot deleted successfully!")
    return redirect(url_for('doctor_dashboard'))



#=============================================================================================================================================================================================================


if __name__ == "__main__":
    # ensure_csv_exists() 
    app.run(debug=True)
